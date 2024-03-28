import sys
import pandas as pd
import openpyxl

#df = openpyxl.load_workbook("oenf1lol.xlsx")
#df1 = df.active
res = ""
#x = df1.max_row
#y = df1.max_column

chal = open("opl.txt", 'r').readlines()
x = len(chal)
y = len(chal[0])

for line in chal:
    line = line.rstrip('\n')
    for letter in line.split(" "):
        if letter != None and letter != " " and letter != "-":
            print(letter)
            res += letter

#for row in range(0, x):
 #   for col in df1.iter_cols(1, y):
  #      letter = col[row].value
   #     if letter != None and letter != " " and letter != "-":
    #        print(letter)
     #       res += letter

print(res)